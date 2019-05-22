import openpyxl
import datetime
from openpyxl import load_workbook
from scale_List import scaleList


wb = load_workbook(filename = "pyExceldata.xlsx")
ws = wb['Sheet1']

for cell in ws['A']:
    if cell.value == "N/A" or cell.value == "No":
        print("skipped")
    else:
        inter = scaleList(cell.value, 10)
        #print(type(inter))
        cell.offset(0,1).value = inter

now = datetime.datetime.now()

nowFormatted = now.strftime("%Y-%m-%d_%H-%M") + "_test.xlsx"

print(type(nowFormatted))
print(nowFormatted)

wb.save(nowFormatted)

#print (sheet_range['A2'].value)
