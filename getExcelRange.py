import openpyxl
import datetime
from openpyxl import load_workbook

wb = load_workbook(filename = "getExcelRange.xlsx")
ws = wb['Sheet2']
desiredRange = []

for cell in ws['A']:
    if cell.value == "N/A" or cell.value == "No":
        print("skipped")
    else:
        desiredRange.append(cell.value)

print(desiredRange)

#wb.save(nowFormatted)


