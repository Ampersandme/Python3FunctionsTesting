#This script is not working as required, vba script to call python script does not work as intended.

import openpyxl
import datetime
from openpyxl import load_workbook

wb = load_workbook(filename = "getExcelRange.xlsx")
ws = wb['Sheet3']
dest_filename = 'test_output.xlsx'

ws['B3'] = "It worked!!"

for cell in ws['A']:
    cell.offset(0, 1).value = cell.value * 2

wb.Save(filename = dest_filename)